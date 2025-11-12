#!/usr/bin/env python3
"""Gerar arquivo ODS de teste com múltiplas abas de pacientes

Cria um arquivo ODS com 3 abas, cada uma com 500+ pacientes,
totalizando 1.600+ pacientes para testar o sistema.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import random

# Criar workbook
wb = Workbook()
wb.remove(wb.active)  # Remove a aba padrão

# Dados para população
sectors = ["Psicologia", "Fisioterapia", "Fonoaudiologia", "Terapia Ocupacional", "Assistência Social"]
modalities = ["Individual", "Grupo", "Dupla"]
routines = ["Semanal", "Quinzenal", "Mensal"]
ubsfs = ["USF Centro", "USF Norte", "USF Sul", "USF Leste", "USF Oeste"]
neighborhoods = ["Centro", "Norte", "Sul", "Leste", "Oeste", "Zona Industrial", "Vila Nova", "Periferia"]
therapists = ["João Silva", "Maria Santos", "Pedro Costa", "Ana Paula", "Carlos Mendes", "Lucia Ferreira", "Roberto Alves", "Fernanda Torres"]
cids = ["F41.1", "M54.0", "F32.9", "F80.0", "F41.0", "M25.9", "F43.1", "F80.1", "F41.2", "M54.1"]

# Cabeçalhos
headers = ["Prontuário", "Nome", "Dias", "Setor", "Modalidade", "Rotina", "UBSF", "Endereço Completo", "Bairro", "Terapeuta", "CID"]

# Criar 3 abas com ~550 pacientes cada
sheet_names = ["Pacientes Jan-Mar", "Pacientes Abr-Jun", "Pacientes Jul-Set"]
start_num = 1

for sheet_idx, sheet_name in enumerate(sheet_names):
    ws = wb.create_sheet(title=sheet_name)
    
    # Adicionar headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Adicionar 550 pacientes por aba
    row = 2
    for i in range(550):
        patient_num = start_num + i
        prontuario = f"P{patient_num:04d}"
        nome = f"Paciente {patient_num:04d}"
        dias = (patient_num * 7) % 365
        setor = sectors[patient_num % len(sectors)]
        modalidade = modalities[patient_num % len(modalities)]
        rotina = routines[patient_num % len(routines)]
        ubsf = ubsfs[patient_num % len(ubsfs)]
        endereco = f"Rua {'ABCDEFGHIJKLMNOPQRST'[(patient_num - 1) % 20]}, {100 + ((patient_num - 1) % 900)}"
        bairro = neighborhoods[patient_num % len(neighborhoods)]
        terapeuta = therapists[patient_num % len(therapists)]
        cid = cids[patient_num % len(cids)]
        
        ws.cell(row=row, column=1, value=prontuario)
        ws.cell(row=row, column=2, value=nome)
        ws.cell(row=row, column=3, value=dias)
        ws.cell(row=row, column=4, value=setor)
        ws.cell(row=row, column=5, value=modalidade)
        ws.cell(row=row, column=6, value=rotina)
        ws.cell(row=row, column=7, value=ubsf)
        ws.cell(row=row, column=8, value=endereco)
        ws.cell(row=row, column=9, value=bairro)
        ws.cell(row=row, column=10, value=terapeuta)
        ws.cell(row=row, column=11, value=cid)
        
        row += 1
    
    start_num += 550
    print(f"✅ Aba '{sheet_name}': 550 pacientes adicionados (P{start_num-550:04d} a P{start_num-1:04d})")

# Salvar como XLSX (compatível com ODS quando aberto)
output_file = "test-multisheet-patients.xlsx"
wb.save(output_file)

print(f"\n✅ Arquivo gerado: {output_file}")
print(f"📊 Total de abas: 3")
print(f"👥 Total de pacientes: 1.650 (550 + 550 + 550)")
print(f"📝 Prontuários: P0001 a P1650")
print(f"\n💡 Abas:")
for sheet_name in sheet_names:
    print(f"   - {sheet_name}")
