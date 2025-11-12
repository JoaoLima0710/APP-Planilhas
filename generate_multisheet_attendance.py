#!/usr/bin/env python3
"""Gerar arquivo XLSX de teste com múltiplas abas de frequência

Cria um arquivo com 3 abas (Novembro, Outubro, Setembro),
cada uma com registros de frequência para os primeiros 100 pacientes.
"""

from openpyxl import Workbook
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
wb.remove(wb.active)  # Remove a aba padrão

# Cabeçalhos
headers = ["Prontuário", "Data Atendimento", "Status"]

# Criar 3 abas por mês
months = [
    ("Novembro", 0),    # Últimos 30 dias
    ("Outubro", 30),    # 30-60 dias atrás
    ("Setembro", 60),   # 60-90 dias atrás
]

total_records = 0

for month_name, days_offset in months:
    ws = wb.create_sheet(title=month_name)
    
    # Adicionar headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Adicionar registros de frequência
    row = 2
    today = datetime.now()
    
    # Para cada paciente (P0001 a P0100)
    for patient_num in range(1, 101):
        prontuario = f"P{patient_num:04d}"
        
        # Cada paciente tem 4-5 atendimentos no mês
        num_records = random.randint(4, 5)
        
        for _ in range(num_records):
            # Data aleatória no mês (30-60 dias atrás para o offset)
            days_ago = days_offset + random.randint(0, 29)
            date = today - timedelta(days=days_ago)
            data_atendimento = date.strftime("%Y-%m-%d")
            
            # 80% de presença, 20% de falta
            status = "P" if random.random() < 0.8 else "F"
            
            ws.cell(row=row, column=1, value=prontuario)
            ws.cell(row=row, column=2, value=data_atendimento)
            ws.cell(row=row, column=3, value=status)
            
            row += 1
            total_records += 1
    
    print(f"✅ Aba '{month_name}': {row - 2} registros de frequência adicionados")

# Salvar como XLSX
output_file = "test-multisheet-attendance.xlsx"
wb.save(output_file)

print(f"\n✅ Arquivo gerado: {output_file}")
print(f"📊 Total de abas: 3 (Novembro, Outubro, Setembro)")
print(f"📝 Total de registros: {total_records}")
print(f"👥 Pacientes: P0001 a P0100")
print(f"📈 Proporção: ~80% presença (P), ~20% falta (F)")
