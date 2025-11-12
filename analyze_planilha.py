#!/usr/bin/env python3
"""
Script para analisar a estrutura da planilha ODS
"""

import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

try:
    import ezodf
except ImportError:
    print("Instalando ezodf...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "ezodf", "lxml"])
    import ezodf

def analyze_ods(file_path):
    """Analisa arquivo ODS"""
    print(f"\n📄 Analisando: {file_path}\n")
    
    try:
        doc = ezodf.opendoc(file_path)
        sheets = doc.sheets
        
        print(f"✅ Total de abas: {len(sheets)}\n")
        
        for sheet_idx, sheet in enumerate(sheets):
            print(f"\n{'='*80}")
            print(f"📑 Aba {sheet_idx + 1}: {sheet.name}")
            print(f"{'='*80}")
            
            # Pega dimensões
            nrows = sheet.nrows()
            ncols = sheet.ncols()
            print(f"Dimensões: {nrows} linhas × {ncols} colunas")
            
            # Analisa headers (primeira linha)
            if nrows > 0:
                headers = []
                print(f"\n📋 Headers (linha 1):")
                for col_idx in range(ncols):
                    cell = sheet[(col_idx, 0)]
                    value = cell.plaintext() if hasattr(cell, 'plaintext') else str(cell.value)
                    headers.append(value)
                    print(f"  Col {col_idx + 1}: '{value}'")
                
                # Analisa primeiras 5 linhas de dados
                print(f"\n📊 Primeiras 5 linhas de dados:")
                for row_idx in range(1, min(6, nrows)):
                    print(f"\n  Linha {row_idx + 1}:")
                    # Mostra só os primeiros 10 valores
                    for col_idx in range(min(10, ncols)):
                        try:
                            cell = sheet[(col_idx, row_idx)]
                            value = cell.plaintext() if hasattr(cell, 'plaintext') else str(cell.value)
                            header = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                            if value and value.strip():  # Só mostra se não vazio
                                print(f"    Col {col_idx + 1} ({header}): {value}")
                        except:
                            pass
                
                # Estatísticas
                print(f"\n📈 Estatísticas:")
                print(f"  Total de linhas (sem header): {nrows - 1}")
                print(f"  Total de colunas: {ncols}")
                
                # Verifica colunas importantes
                print(f"\n🔍 Colunas Importantes Detectadas:")
                important_cols = ["PRONTUÁRIO", "NOMES", "SETOR", "UBSF", "TERAPEUTA", "DIAS", "FALTAS"]
                for important in important_cols:
                    found = any(important.upper() in h.upper() for h in headers)
                    status = "✅ ENCONTRADA" if found else "❌ NÃO ENCONTRADA"
                    print(f"  {important}: {status}")
                    if found:
                        col_name = next(h for h in headers if important.upper() in h.upper())
                        print(f"    → Nome exato na planilha: '{col_name}'")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao ler ODS: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("🚀 Analisador de Planilhas ODS\n")
        print("Uso: python analyze_planilha.py <arquivo.ods>\n")
        print("Exemplo:")
        print("  python analyze_planilha.py planilha.ods")
        print("  python analyze_planilha.py C:\\Users\\Joao\\Desktop\\minha_planilha.ods\n")
        sys.exit(1)
    
    file_path = sys.argv[1]
    path = Path(file_path)
    
    if not path.exists():
        print(f"❌ Arquivo não encontrado: {file_path}")
        sys.exit(1)
    
    success = analyze_ods(file_path)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
