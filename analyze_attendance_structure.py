#!/usr/bin/env python3
"""
Analisar estrutura do arquivo de presença (ODS/XLSX) para debug
"""
import openpyxl
import sys
from pathlib import Path

# Tentar com arquivo padrão
attendance_file = Path("C:/Users/Joao/Downloads/SUL-NOVEMBRO.ods")

# Se não existir, procurar por outros arquivos .ods
if not attendance_file.exists():
    ods_files = list(Path("C:/Users/Joao/Downloads").glob("*.ods"))
    if ods_files:
        attendance_file = ods_files[0]
        print(f"Arquivo não encontrado, usando: {attendance_file}")
    else:
        print("Nenhum arquivo .ods encontrado")
        sys.exit(1)

print(f"Analisando: {attendance_file}")
print("=" * 80)

try:
    # Tentar abrir com openpyxl (funciona com XLSX e alguns ODS)
    wb = openpyxl.load_workbook(attendance_file, data_only=True)
    sheets = wb.sheetnames
    
except Exception as e:
    print(f"Erro com openpyxl: {e}")
    print("\nTentando com ezodf...")
    
    try:
        from odfpy import ezodf
        doc = ezodf.opendoc(str(attendance_file))
        sheets = [sheet.name for sheet in doc.sheets]
        wb = doc
    except Exception as e2:
        print(f"Erro com ezodf: {e2}")
        print("\nTentando com pandas + odfpy...")
        
        try:
            import pandas as pd
            xls = pd.ExcelFile(str(attendance_file), engine='odf')
            sheets = xls.sheet_names
            wb = xls
        except Exception as e3:
            print(f"Erro com pandas: {e3}")
            sys.exit(1)

print(f"\n📊 Arquivo tem {len(sheets)} abas:\n")

for sheet_name in sheets:
    print(f"  • {sheet_name}")

print("\n" + "=" * 80)
print("\n🔍 Analisando primeira aba: {0}\n".format(sheets[0]))

try:
    # Carregar primeira aba
    if hasattr(wb, 'sheet_by_name'):  # xlrd style
        ws = wb.sheet_by_name(sheets[0])
        rows = []
        for row_idx in range(min(15, ws.nrows)):
            row = []
            for col_idx in range(min(15, ws.ncols)):
                cell = ws.cell_value(row_idx, col_idx)
                row.append(str(cell) if cell is not None else "")
            rows.append(row)
    else:
        # openpyxl style
        ws = wb[sheets[0]]
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=15, max_col=15)):
            rows.append([str(cell) if cell is not None else "" for cell in row])

    # Exibir primeiras 10 linhas e 10 colunas
    print("Primeiras 10 linhas x 10 colunas:\n")
    
    # Header com números de coluna
    print("    ", end="")
    for col in range(min(10, len(rows[0]) if rows else 0)):
        print(f"Col{col:2d}  ", end="")
    print()
    print("    " + "-" * 60)
    
    for row_idx, row in enumerate(rows[:10]):
        print(f"L{row_idx+1:2d} | ", end="")
        for col_idx in range(min(10, len(row))):
            cell_val = str(row[col_idx])[:5].ljust(5)
            print(f"{cell_val}  ", end="")
        print()
    
    print("\n" + "=" * 80)
    
    # Procurar por prontuários (números de 4-8 dígitos)
    print("\n🔎 Procurando por prontuários (números 4-8 dígitos)...\n")
    
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).strip()
            if len(cell_str) >= 4 and len(cell_str) <= 8 and cell_str.isdigit():
                print(f"  ✓ Encontrado prontuário '{cell_str}' em L{row_idx+1} Col{col_idx+1}")
    
    # Procurar por P/F (status de presença)
    print("\n🔎 Procurando por status (P ou F)...\n")
    
    status_found = False
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).strip().upper()
            if cell_str in ['P', 'F']:
                print(f"  ✓ Encontrado status '{cell_str}' em L{row_idx+1} Col{col_idx+1}")
                status_found = True
    
    if not status_found:
        print("  ✗ Nenhum status P/F encontrado")
    
    print("\n" + "=" * 80)
    
except Exception as e:
    print(f"Erro ao analisar: {e}")
    import traceback
    traceback.print_exc()

print("\n✅ Análise concluída")
