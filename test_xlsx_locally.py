"""
Script para testar o processamento de XLSX localmente
Simula o que a Edge Function faz
"""
import openpyxl

# Ler o arquivo de teste
file_path = "test-simple-pacientes.xlsx"
wb = openpyxl.load_workbook(file_path)

print(f"📂 Arquivo: {file_path}")
print(f"📋 Abas: {wb.sheetnames}\n")

all_rows = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"🔍 Processando aba: {sheet_name}")
    
    # Extrair header (primeira linha)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"   Headers: {headers}")
    
    # Extrair dados
    rows_in_sheet = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:  # Parar em linhas vazias
            break
        
        # Montar objeto com headers
        row_data = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_data[header.upper() if header else f"COL_{i}"] = row[i]
        
        all_rows.append(row_data)
        rows_in_sheet += 1
        print(f"   Linha {row_idx}: {row_data}")
    
    print(f"   ✅ {rows_in_sheet} linhas processadas\n")

print(f"\n📊 RESUMO:")
print(f"   Total de linhas: {len(all_rows)}")
print(f"   Planilha deve mostrar: {len(all_rows)} pacientes")

# Verificar se as colunas necessárias estão presentes
required_cols = ["PRONTUÁRIO", "NOME", "SETOR"]
print(f"\n✅ Validação:")
for row in all_rows[:1]:
    for col in required_cols:
        if col in row:
            print(f"   ✓ '{col}' encontrado")
        else:
            print(f"   ✗ '{col}' NÃO encontrado")
